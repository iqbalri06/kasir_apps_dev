from django.urls import reverse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import Group
from django.db import models
from django.db.models import Sum, Count, Q
from django.http import JsonResponse, HttpResponse, Http404
from django.template.loader import render_to_string, get_template
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.utils.crypto import get_random_string
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.generic import UpdateView
from django.db.models.functions import ExtractMonth, Coalesce

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A6, landscape
from reportlab.lib.units import mm, inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader  # Add this import at the top with other imports
from reportlab.lib.colors import Color, HexColor

import json
import qrcode
import random
import string
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from .models import (
    Product, CartItem, Sale, Member, User, Payment, 
    SaleDetail, StockMovement, CustomerType, Receipt, Role
)
from .forms import ProductForm, UserForm, AddUserForm
from .utils import generate_receipt_pdf
from .email_utils import send_otp_email as send_otp_mail

import uuid  # Add this import

def is_admin(user):
    return user.role and user.role.name == 'Admin'

def is_cashier(user):
    return user.role and user.role.name == 'Cashier'
def dashboard(request):
    if request.user.role and request.user.role.name == 'Admin':
        return admin_dashboard(request)
    else:
        return cashier_dashboard(request)

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    try:
        now = timezone.localtime(timezone.now())
        today = now.date()
        yesterday = today - timezone.timedelta(days=1)
        first_day = today.replace(day=1)
        
        # Last month's date range
        last_month_end = first_day - timezone.timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        
        # Last week's date
        last_week = today - timezone.timedelta(days=7)
        
        # Debug prints
        print(f"Debug - Time info:")
        print(f"Server timezone: {timezone.get_current_timezone()}")
        print(f"Current time: {now}")
        print(f"Today date: {today}")
        print(f"Yesterday date: {yesterday}")
        print(f"First day of month: {first_day}")
        print(f"Last month start: {last_month_start}")
        print(f"Last month end: {last_month_end}")

        # Today's sales
        today_sales_qs = Sale.objects.filter(
            created_at__date=today
        )
        today_sales = today_sales_qs.aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        # Yesterday's sales for comparison
        yesterday_sales = Sale.objects.filter(
            created_at__date=yesterday
        ).aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        # Calculate daily sales change percentage
        daily_sales_change = 0
        if yesterday_sales > 0:
            daily_sales_change = ((float(today_sales) - float(yesterday_sales)) / float(yesterday_sales)) * 100
        
        # Monthly sales (current month)
        monthly_sales_qs = Sale.objects.filter(
            created_at__date__range=[first_day, today]
        )
        monthly_sales = monthly_sales_qs.aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        # Last month's sales for comparison
        last_month_sales = Sale.objects.filter(
            created_at__date__range=[last_month_start, last_month_end]
        ).aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        # Calculate monthly sales change percentage
        monthly_sales_change = 0
        if last_month_sales > 0:
            monthly_sales_change = ((float(monthly_sales) - float(last_month_sales)) / float(last_month_sales)) * 100
        
        # Daily transactions count
        daily_transactions = today_sales_qs.count()
        
        # Yesterday's transactions count
        yesterday_transactions = Sale.objects.filter(
            created_at__date=yesterday
        ).count()
        
        # Calculate daily transactions change percentage
        daily_transactions_change = 0
        if yesterday_transactions > 0:
            daily_transactions_change = ((daily_transactions - yesterday_transactions) / yesterday_transactions) * 100
        
        # Monthly transactions
        monthly_transactions = monthly_sales_qs.count()
        
        # Current members count
        members_count = Member.objects.count()
        
        # Members count from a week ago
        members_last_week = Member.objects.filter(
            created_at__lt=last_week
        ).count() 
        
        # Calculate members count change percentage
        members_count_change = 0
        if members_last_week > 0:
            members_count_change = ((members_count - members_last_week) / members_last_week) * 100

        # Get monthly data for chart
        monthly_data = []
        for month in range(1, 13):
            month_total = Sale.objects.filter(
                created_at__year=today.year,
                created_at__month=month
            ).aggregate(
                total=Sum('total_price')
            )['total'] or 0
            monthly_data.append(float(month_total))

        # Get daily sales data for last 7 days
        from datetime import timedelta
        last_7_days = [(today - timedelta(days=x)) for x in range(6, -1, -1)]
        daily_sales_data = []
        daily_labels = []
        
        for day in last_7_days:
            daily_total = Sale.objects.filter(
                created_at__date=day
            ).aggregate(
                total=Sum('total_price')
            )['total'] or 0
            daily_sales_data.append(float(daily_total))
            daily_labels.append(day.strftime('%d/%m'))

        # Product share data - UPDATED
        # First get total number of products sold
        total_products_sold = SaleDetail.objects.filter(
            sale__created_at__year=today.year
        ).count()

        # Then get top products with percentages
        top_products = SaleDetail.objects.filter(
            sale__created_at__year=today.year
        ).values(
            'product__name'
        ).annotate(
            total_sold=Count('id')
        ).order_by('-total_sold')[:5]

        # Calculate percentages
        pie_data = {}
        for item in top_products:
            percentage = (item['total_sold'] / total_products_sold * 100) if total_products_sold > 0 else 0
            pie_data[item['product__name']] = round(percentage, 1)  # Round to 1 decimal place

        print(f"Debug - Total products sold: {total_products_sold}")
        print(f"Debug - Product percentages: {pie_data}")

        context = {
            'today_sales': float(today_sales),
            'total_sales': float(monthly_sales),  # Renamed from monthly_sales for clarity
            'members_count': members_count,
            'daily_transactions_count': daily_transactions,
            'monthly_transactions_count': monthly_transactions,
            'monthly_sales_data': monthly_data,
            'daily_sales_data': daily_sales_data,
            'daily_labels': daily_labels,
            'product_share_labels': list(pie_data.keys()),
            'product_share_data': list(pie_data.values()),
            'today': today,
            # Add the new percentage changes to context
            'daily_sales_change': daily_sales_change,
            'monthly_sales_change': monthly_sales_change,
            'daily_transactions_change': daily_transactions_change,
            'members_count_change': members_count_change,
        }
        
        return render(request, 'dashboard_admin.html', context)
    
    except Exception as e:
        print(f"Error in admin_dashboard: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise

@login_required
@user_passes_test(is_cashier)

def cashier_dashboard(request):
    try:
        now = timezone.localtime(timezone.now())
        today = now.date()
        
        print(f"Cashier dashboard - date/time check:")
        print(f"Server timezone: {timezone.get_current_timezone()}")
        print(f"Current time: {now}")
        print(f"Today date: {today}")
        print(f"Current user: {request.user.id}, {request.user.username}")

        # Get today's sales - removed user filter to show all transactions
        today_sales_qs = Sale.objects.filter(
            created_at__date=today
        )
        
        # Get sales total
        today_sales = today_sales_qs.aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        # Get transaction count
        transactions_count = today_sales_qs.count()

        # Calculate average transaction value
        average_transaction = float(today_sales) / transactions_count if transactions_count > 0 else 0

        # Get today's sales with pagination - removed user filter
        sales = Sale.objects.filter(
            created_at__date=today  # Only show today's transactions
        ).select_related(
            'member', 'payment', 'user'  # Added 'user' to show who processed each transaction
        ).order_by('-created_at')

        if request.GET.get('search'):
            search_query = request.GET.get('search')
            sales = sales.filter(
                Q(transaction_number__icontains=search_query) |
                Q(member__name__icontains=search_query) |
                Q(member__phone__icontains=search_query) |
                Q(user__username__icontains=search_query)  # Added search by cashier username
            )

        paginator = Paginator(sales, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        context = {
            'today_sales': float(today_sales),
            'transactions_count': transactions_count,
            'average_transaction': float(average_transaction),
            'sales': page_obj,
            'search_query': request.GET.get('search', ''),
            'today': today,
        }

        return render(request, 'dashboard_cashier.html', context)

    except Exception as e:
        print(f"Error in cashier_dashboard: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise

@login_required
def product_list(request):
    search_query = request.GET.get('search', '')
    if search_query:
        products = Product.objects.filter(name__icontains=search_query)
    else:
        products = Product.objects.all()
    # Cast to list to evaluate the queryset and catch any database errors
    try:
        products.exists()  # Test the query
    except Exception as e:
        print(f"Database error: {str(e)}")
        products = []
    return render(request, 'products/product_list.html', {'products': products})

@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'Product {product.name} created successfully')
            return redirect('simple_shop:products')
    else:
        form = ProductForm()
    return render(request, 'products/add_product.html', {'form': form})

@login_required
def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f'Product {product.name} updated successfully')
            return redirect('simple_shop:products')
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/product_detail.html', {'form': form, 'product': product})

class ProductUpdateView(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_detail.html'
    
    def get_initial(self):
        initial = super().get_initial()
        # Format harga tanpa menambah digit
        if self.object:
            initial['price'] = str(self.object.price)
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.object
        # Kirim harga asli tanpa format dan desimal
        context['original_price'] = int(self.object.price)
        return context

def cart(request):
    cart_items = CartItem.objects.all()
    total = sum(item.total_price() for item in cart_items)
    return render(request, 'cart.html', {'cart_items': cart_items, 'total': total})

@login_required
def add_to_cart(request, product_id):
    if request.method == 'POST':
        try:
            product = get_object_or_404(Product, id=product_id)
            data = json.loads(request.body)
            quantity = int(data.get('quantity', 1))
            
            # Check if product already in cart
            cart_item = CartItem.objects.filter(product=product).first()
            
            if cart_item:
                # Update existing cart item
                cart_item.quantity += quantity
                cart_item.save()
            else:
                # Create new cart item
                cart_item = CartItem.objects.create(
                    product=product,
                    quantity=quantity
                )
            
            # Calculate new cart total
            cart_total = sum(item.total_price() for item in CartItem.objects.all())
            
            return JsonResponse({
                'success': True,
                'message': f'{product.name} added to cart',
                'cart_total': cart_total
            })
            
        except Exception as e:
            print(f"Error adding to cart: {str(e)}")  # Debug print
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    })

@login_required
def add_to_cart_ajax(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        cart_item, created = CartItem.objects.get_or_create(
            product=product,
            defaults={'quantity': 1}
        )
        
        if not created:
            cart_item.quantity += 1
            cart_item.save()
        
        cart_total = sum(item.total_price() for item in CartItem.objects.all())
        
        return JsonResponse({
            'success': True,
            'message': f'{product.name} added to cart',
            'cart_total': cart_total
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def get_cart_ajax(request):
    cart_items = CartItem.objects.select_related('product').all()
    total = sum(item.total_price() for item in cart_items)
    
    cart_html = render_to_string('cashier/cart_items.html', {
        'cart_items': cart_items,
        'total': total
    }, request=request)
    
    return JsonResponse({
        'success': True,
        'cart_html': cart_html,
        'total': f'Rp {total:,.0f}'.replace(',', '.')
    })

@login_required
def remove_from_cart_ajax(request, item_id):
    try:
        item = get_object_or_404(CartItem, id=item_id)
        product_id = item.product.id
        quantity = item.quantity
        item.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Item removed from cart',
            'product_id': product_id,
            'quantity': quantity,
            'cart_total': sum(item.total_price() for item in CartItem.objects.all())
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
def clear_cart(request):
    if request.method == 'POST':
        try:
            CartItem.objects.all().delete()
            
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Keranjang berhasil dikosongkan'
                })
            
            # Handle regular request
            messages.success(request, 'Keranjang berhasil dikosongkan')
            return redirect('simple_shop:cashier')
            
        except Exception as e:
            # Handle errors for AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                })
            
            # Handle errors for regular request
            messages.error(request, f'Gagal mengosongkan keranjang: {str(e)}')
            return redirect('simple_shop:cashier')
    
    # Fallback for non-POST requests
    return redirect('simple_shop:cashier')

def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product.delete()
    messages.success(request, f'Produk {product.name} berhasil dihapus')
    return redirect('simple_shop:products')

def delete_cart_item(request, item_id):
    item = get_object_or_404(CartItem, id=item_id)
    item.delete()
    return redirect('cart')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Clear any existing messages
        storage = messages.get_messages(request)
        storage.used = True
        
        try:
            # Try to get user first to check if account exists
            User = get_user_model()
            user = User.objects.filter(username=username).first()
            
            if user is None:
                # User doesn't exist
                return render(request, 'registration/login.html', {
                    'error_message': 'Username tidak ditemukan.',
                    'error_type': 'invalid-credentials',
                    'username': username
                })
                
            if not user.is_active:
                # Account is disabled
                return render(request, 'registration/login.html', {
                    'error_message': 'Akun telah dinonaktifkan. Silakan hubungi administrator.',
                    'error_type': 'account-disabled',
                    'username': username
                })
            
            # Try to authenticate
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                next_url = request.GET.get('next', reverse('simple_shop:dashboard'))
                return redirect(next_url)
            else:
                # Wrong password
                return render(request, 'registration/login.html', {
                    'error_message': 'Password salah. Silakan coba lagi.',
                    'error_type': 'invalid-credentials',
                    'username': username
                })
                
        except Exception as e:
            print(f"Login error: {str(e)}")
            return render(request, 'registration/login.html', {
                'error_message': 'Terjadi kesalahan. Silakan coba lagi.',
                'error_type': 'error',
                'username': username
            })
    
    return render(request, 'registration/login.html')

def logout_view(request):
    logout(request)
    return redirect('simple_shop:login')

@login_required
@user_passes_test(is_admin)
def user_list(request):
    # Base query
    users = User.objects.select_related('role').all()
    
    # Filter pencarian
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) | 
            Q(email__icontains=search_query)
        )
    
    # Filter berdasarkan role
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role__name__iexact=role_filter)
    
    # Filter berdasarkan status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    # Hitung statistik untuk tampilan
    total_users = User.objects.count()
    active_count = User.objects.filter(is_active=True).count()
    inactive_count = User.objects.filter(is_active=False).count()  # Fixed: using is_active=False instead of is_active()
    admin_count = User.objects.filter(role__name='Admin').count()  # Added missing admin_count
    
    # Add pagination
    paginator = Paginator(users, 12)  # Show 12 users per page
    page = request.GET.get('page', 1)
    
    try:
        users = paginator.page(page)
    except:
        users = paginator.page(1)
    
    context = {
        'users': users,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'admin_count': admin_count,
        'total_users': total_users,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'users/user_list.html', context)

@login_required
@user_passes_test(is_admin)
def add_user(request):
    if request.method == 'POST':
        form = AddUserForm(request.POST)  # Use the AddUserForm for new users
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = True  # Set all new users as staff automatically
            user.save()
            messages.success(request, f'User {user.username} berhasil dibuat dengan role {user.role.name}')
            return redirect('simple_shop:users')
    else:
        form = AddUserForm()  # Use the AddUserForm for new users

    context = {
        'form': form,
        'roles': Role.objects.all().order_by('name'),
        'title': 'Tambah User'
    }
    return render(request, 'users/add_user.html', context)

@login_required
@user_passes_test(is_admin)
def user_detail(request, pk):  # Changed parameter from user_id to pk
    User = get_user_model()
    user_obj = get_object_or_404(User, pk=pk)
    context = {
        'user_detail': user_obj,
    }
    return render(request, 'users/user_detail.html', context)

@login_required
@user_passes_test(is_admin)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully')
    return redirect('simple_shop:users')  # Add namespace

@login_required
@user_passes_test(is_admin)
def sale_list(request):
    try:
        # Base queryset with related fields
        sales = Sale.objects.select_related(
            'user', 'member', 'payment'
        ).order_by('-created_at')

        # Search functionality
        search_query = request.GET.get('search', '')
        if search_query:
            sales = sales.filter(
                Q(id__icontains=search_query) |  # Search by transaction ID
                Q(member__name__icontains=search_query) |
                Q(member__phone__icontains=search_query) |
                Q(user__username__icontains=search_query)
            )

        # Date filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        payment_method = request.GET.get('payment_method')

        if date_from:
            sales = sales.filter(created_at__date__gte=date_from)
        if date_to:
            sales = sales.filter(created_at__date__lte=date_to)
        if payment_method:
            sales = sales.filter(payment__method=payment_method)

        # Export functionality if needed
        if request.GET.get('export') == 'excel':
            return export_sales(request)

        # Debug print to check data
        print(f"Total sales before pagination: {sales.count()}")

        # Pagination
        paginator = Paginator(sales, 10)  # Show 10 sales per page
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        print(f"Page {page_number} has {len(page_obj)} items")

        context = {
            'sales': page_obj,
            'search_query': search_query,
            'date_from': date_from,
            'date_to': date_to,
            'payment_method': payment_method,
            'payment_methods': Payment.objects.values_list('method', flat=True).distinct(),
            'total_sales': sales.count()
        }
        
        return render(request, 'sales/sale_list.html', context)

    except Exception as e:
        print(f"Error in sale_list view: {str(e)}")
        import traceback
        print(traceback.format_exc())
        messages.error(request, f"Error retrieving sales data: {str(e)}")
        return render(request, 'sales/sale_list.html', {'sales': []})

@login_required
def new_sale(request):
    cart_items = CartItem.objects.all()
    if not cart_items:
        messages.warning(request, 'Cart is empty')
        return redirect('simple_shop:products')  # Add namespace

    if request.method == 'POST':
        payment_id = request.POST.get('payment')
        member_id = request.POST.get('member')
        payment = get_object_or_404(Payment, id=payment_id)
        member = Member.objects.filter(id=member_id).first()
        
        total_price = sum(item.total_price() for item in cart_items)
        
        sale = Sale.objects.create(
            user=request.user,
            member=member,
            payment=payment,
            total_price=total_price,
            status='completed'  # Add status here too
        )
        
        for item in cart_items:
            SaleDetail.objects.create(
                sale=sale,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price,
                subtotal=item.total_price()
            )
            
            # Update stock
            StockMovement.objects.create(
                product=item.product,
                type='OUT',
                quantity=item.quantity
            )
        
        cart_items.delete()
        messages.success(request, 'Sale completed successfully')
        return redirect('simple_shop:sale_detail', sale_id=sale.id)  # Add namespace
    
    members = Member.objects.all()
    payments = Payment.objects.all()
    total = sum(item.total_price() for item in cart_items)

    context = {
        'cart_items': cart_items,
        'members': members,
        'payments': payments,        
        'total': total,
    }
    return render(request, 'sales/new_sale.html', context)

@login_required
def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale.objects.select_related(
        'user', 'member', 'payment'
    ).prefetch_related('details__product'), id=sale_id)

    context = {
        'sale': sale,
        'details': sale.details.all(),
    }
    return render(request, 'sales/sale_detail.html', context)

@login_required
@user_passes_test(is_cashier)
def cashier(request):
    products = Product.objects.select_related('category').all()
    cart_items = CartItem.objects.select_related('product').all()
    payments = Payment.objects.all().order_by('method')  # Add order_by to ensure consistent ordering
    total = sum(item.total_price() for item in cart_items)
    members = Member.objects.all()
    
    context = {
        'products': products,
        'cart_items': cart_items,
        'total': total,
        'payments': payments,
        'members': members,
    }
    return render(request, 'cashier/index.html', context)

@login_required
def search_member(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            phone = data.get('phone')
            name = data.get('name')
            
            # Try to find existing member
            member = Member.objects.filter(phone=phone).first()
            if member:
                return JsonResponse({
                    'success': True,
                    'member': {
                        'id': str(member.id),
                        'name': member.name,
                        'points': member.points,
                        'phone': member.phone
                    }
                })
            
            # Create new member
            if name:
                member = Member.objects.create(
                    phone=phone,
                    name=name,
                    customer_type=CustomerType.objects.get(name='Regular')
                )
                return JsonResponse({
                    'success': True,
                    'member': {
                        'id': str(member.id),
                        'name': member.name,
                        'points': 0,
                        'phone': member.phone
                    },
                    'message': f'Member {name} berhasil didaftarkan'
                })
            
            return JsonResponse({
                'success': False,
                'message': 'Nama member harus diisi'
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal mendaftarkan member: {str(e)}'
            })
    
    # Handle GET request for member search
    phone = request.GET.get('phone')
    try:
        member = Member.objects.get(phone=phone)
        return JsonResponse({
            'success': True,
            'member': {
                'id': str(member.id),
                'name': member.name,
                'points': member.points,
                'phone': member.phone
            }
        })
    except Member.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Member tidak ditemukan'
        })

@login_required
def register_member(request):
    if request.method == 'POST':
        try:
            phone = request.POST.get('phone')
            name = request.POST.get('name')
            
            print(f"Received member registration: {phone=}, {name=}")  # Debug print
            
            if not phone or not name:
                return JsonResponse({
                    'success': False,
                    'message': 'Nomor telepon dan nama harus diisi'
                })
            
            # Create new member
            member = Member.objects.create(
                phone=phone,
                name=name,
                customer_type=CustomerType.objects.get(name='Regular'),
                points=0
            )
            
            return JsonResponse({
                'success': True,
                'member': {
                    'id': str(member.id),
                    'name': member.name,
                    'points': member.points,
                    'phone': member.phone
                },
                'message': f'Member {name} berhasil didaftarkan'
            })

        except Exception as e:
            print(f"Error registering member: {str(e)}")  # Debug print
            return JsonResponse({
                'success': False,
                'message': f'Gagal mendaftarkan member: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def checkout(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            payment_id = data.get('payment_id')
            member_phone = data.get('member_phone')
            points_to_use = int(data.get('points_to_use', 0))
            
            # Add cash amount handling
            cash_amount = data.get('cash_amount', '0')
            
            cart_items = CartItem.objects.select_related('product').all()
            if not cart_items.exists():
                return JsonResponse({'success': False, 'message': 'Keranjang kosong'})
            
            total_price = sum(item.total_price() for item in cart_items)
            discount_amount = 0
            points_earned = 0

            # Handle member and points
            member = None
            if member_phone:
                member, created = Member.objects.get_or_create(
                    phone=member_phone,
                    defaults={
                        'name': f"Member {member_phone}",
                        'customer_type': CustomerType.objects.get(name='Regular')
                    }
                )
                
                if points_to_use > 0:
                    discount_amount = member.use_points(points_to_use)
                
                # Calculate final price after discount
                final_price = total_price - discount_amount
                points_earned = member.add_points(final_price)
            else:
                final_price = total_price

            payment = get_object_or_404(Payment, id=payment_id)

            # Store payment details
            payment_details = {}
            if payment.method.lower() == 'cash':
                payment_details['cash_amount'] = cash_amount

            sale = Sale.objects.create(
                user=request.user,
                member=member,
                payment=payment,
                total_price=final_price,
                points_earned=points_earned,
                points_used=points_to_use,
                discount_amount=discount_amount,
                payment_details=payment_details,
                status='completed'  # Set default status to completed
            )
            
            # Process cart items
            for item in cart_items:
                SaleDetail.objects.create(
                    sale=sale,
                    product=item.product,
                    quantity=item.quantity,
                    price=item.product.price,
                    subtotal=item.total_price()
                )
                
                StockMovement.objects.create(
                    product=item.product,
                    type='OUT',
                    quantity=item.quantity,
                )
            
            # Create receipt
            Receipt.objects.create(
                sale=sale,
                printed_by=request.user
            )
            
            # Clear cart after successful transaction
            cart_items.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Penjualan berhasil diselesaikan',
                'sale_id': str(sale.id),
                'points_earned': points_earned,
                'points_used': points_to_use,
                'discount_amount': discount_amount,
                'redirect_url': reverse('simple_shop:print_receipt', args=[sale.id])
            })
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())  # Debug print
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Metode request tidak valid'})

@login_required
def print_receipt(request, sale_id):
    sale = get_object_or_404(Sale.objects.select_related(
        'user', 'member', 'payment'
    ).prefetch_related('details__product'), id=sale_id)

    pdf = generate_receipt_pdf(sale)
    # Format: STRUK_DDMMYYYY_HHMMSS_TRANSACTIONID
    filename = f"STRUK_{sale.created_at.strftime('%d%m%Y_%H%M%S')}_{str(sale.id)[:8]}.pdf"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Transfer-Encoding'] = 'binary'
    response['Accept-Ranges'] = 'bytes'
    
    # Write PDF content
    response.write(pdf.getvalue())
    pdf.close()
    
    return response

def get_filtered_sales(request):
    """Helper function to get filtered sales data"""
    sales = Sale.objects.select_related(
        'user', 'member', 'payment'
    ).prefetch_related(
        'details__product'  # Changed from saleitems__product to details__product
    ).order_by('-created_at')

    # Apply search filter
    search_query = request.GET.get('search')
    if search_query:
        sales = sales.filter(
            Q(transaction_number__icontains=search_query) |
            Q(member__name__icontains=search_query) |
            Q(member__phone__icontains=search_query)
        )

    # Apply date filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        sales = sales.filter(created_at__date__gte=date_from)
    if date_to:
        sales = sales.filter(created_at__date__lte=date_to)

    # Apply payment method filter if provided
    payment_method = request.GET.get('payment_method')
    if payment_method:
        sales = sales.filter(payment__method=payment_method)

    return sales

@login_required
def export_sales(request):
    try:
        # Check if request is from cashier dashboard
        is_daily_report = request.GET.get('report_type') == 'daily'
        
        # Get filtered sales using the helper function
        sales = get_filtered_sales(request)
        
        # If it's a daily report from cashier dashboard, filter for today only
        if is_daily_report:
            today = timezone.localtime(timezone.now()).date()
            sales = sales.filter(created_at__date=today)

        # Create workbook with a nice style
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Penjualan"
        
        # Create styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        subtitle_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        number_font = Font(name='Arial', size=10)
        date_font = Font(name='Arial', size=10)
        total_font = Font(name='Arial', size=11, bold=True)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create title
        title = "LAPORAN PENJUALAN HARIAN" if is_daily_report else "LAPORAN PENJUALAN"
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:H1')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add date range subtitle
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        date_range = "Semua Transaksi"
        if date_from and date_to:
            date_range = f"Periode: {date_from} s/d {date_to}"
        elif date_from:
            date_range = f"Mulai dari: {date_from}"
        elif date_to:
            date_range = f"Sampai dengan: {date_to}"
            
        subtitle_cell = ws.cell(row=2, column=1, value=date_range)
        subtitle_cell.font = subtitle_font
        ws.merge_cells('A2:H2')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add spacing
        ws.row_dimensions[3].height = 10
        
        # Define headers
        headers = [
            'Tanggal', 'No Transaksi', 'Customer', 'Item', 'Qty', 
            'Harga Satuan', 'Total Harga', 'Total Transaksi'
        ]
        
        # Write headers with style
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process sales data
        row = 5
        grand_total = 0
        total_items = 0
        unique_transactions = 0
        
        for sale in sales:
            # Get all items for this sale
            sale_items = sale.details.all()
            first_item = True
            
            # Track unique transactions
            unique_transactions += 1
            # Add to grand total
            grand_total += float(sale.total_price)

            for item in sale_items:
                # Count total items
                total_items += item.quantity
                
                # Common sale data (only on first item)
                if first_item:
                    # Convert timezone-aware datetime to naive datetime for Excel
                    naive_datetime = sale.created_at.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)
                    
                    # Date cell with naive datetime
                    date_cell = ws.cell(row=row, column=1, value=naive_datetime)
                    date_cell.font = date_font
                    date_cell.alignment = Alignment(horizontal='center', vertical='center')
                    date_cell.number_format = 'dd-mmm-yyyy h:mm'
                    
                    # Transaction number
                    trans_cell = ws.cell(row=row, column=2, value=sale.transaction_number)
                    trans_cell.font = normal_font
                    trans_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Customer name
                    cust_cell = ws.cell(row=row, column=3, value=sale.member.name if sale.member else "Non Member")
                    cust_cell.font = normal_font
                    cust_cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Transaction total
                    total_cell = ws.cell(row=row, column=8, value=float(sale.total_price))
                    total_cell.font = number_font
                    total_cell.alignment = Alignment(horizontal='right', vertical='center')
                    total_cell.number_format = '#,##0.00'
                    
                    # Merge cells if there are multiple items
                    if len(sale_items) > 1:
                        for col in [1, 2, 3, 8]:
                            ws.merge_cells(
                                start_row=row,
                                start_column=col,
                                end_row=row + len(sale_items) - 1,
                                end_column=col
                            )
                    
                    first_item = False
                
                # Product name
                name_cell = ws.cell(row=row, column=4, value=item.product.name)
                name_cell.font = normal_font
                name_cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Quantity
                qty_cell = ws.cell(row=row, column=5, value=item.quantity)
                qty_cell.font = normal_font
                qty_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Unit price
                price_cell = ws.cell(row=row, column=6, value=float(item.price))
                price_cell.font = normal_font
                price_cell.alignment = Alignment(horizontal='right', vertical='center')
                price_cell.number_format = '#,##0.00'
                
                # Subtotal
                subtotal_cell = ws.cell(row=row, column=7, value=float(item.subtotal))
                subtotal_cell.font = normal_font
                subtotal_cell.alignment = Alignment(horizontal='right', vertical='center')
                subtotal_cell.number_format = '#,##0.00'
                
                # Add borders to all cells in the row
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1
        
        # Add spacer row
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        row += 1
        
        # Add summary section
        summary_labels = [
            "Ringkasan Laporan", "", "", "", "",
            "Total Item Terjual:", 
            "Total Penjualan:", 
            "Grand Total:"
        ]
        
        # Apply subtitle style to summary header
        summary_header_cell = ws.cell(row=row, column=1, value=summary_labels[0])
        summary_header_cell.font = subtitle_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # Transaction count summary
        count_label_cell = ws.cell(row=row, column=1, value="Jumlah Transaksi:")
        count_label_cell.font = normal_font
        count_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        
        count_value_cell = ws.cell(row=row, column=6, value=unique_transactions)
        count_value_cell.font = normal_font
        count_value_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Items summary
        items_label_cell = ws.cell(row=row, column=1, value=summary_labels[5])
        items_label_cell.font = normal_font
        items_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        
        items_value_cell = ws.cell(row=row, column=6, value=total_items)
        items_value_cell.font = normal_font
        items_value_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Sales total summary with bold formatting
        total_label_cell = ws.cell(row=row, column=1, value=summary_labels[6])
        total_label_cell.font = total_font
        total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'A{row}:E{row}')
        
        total_value_cell = ws.cell(row=row, column=6, value="")
        total_value_cell.font = normal_font
        
        grand_total_cell = ws.cell(row=row, column=8, value=grand_total)
        grand_total_cell.font = total_font
        grand_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        grand_total_cell.number_format = '#,##0.00'
        grand_total_cell.border = Border(
            bottom=Side(style='double'),
            top=Side(style='thin')
        )
        row += 2
        
        # Add export date
        current_time = timezone.now().astimezone(timezone.get_current_timezone())
        naive_export_time = current_time.replace(tzinfo=None)
        export_date_cell = ws.cell(row=row, column=1, 
            value=f"Diekspor pada: {naive_export_time.strftime('%d-%m-%Y %H:%M:%S')}")
        export_date_cell.font = Font(size=8, italic=True)
        export_date_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'A{row}:D{row}')
        
        # Add user info
        user_cell = ws.cell(row=row, column=6, value=f"Oleh: {request.user.username}")
        user_cell.font = Font(size=8, italic=True)
        user_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'F{row}:H{row}')

        # Adjust column widths
        column_widths = [20, 18, 25, 35, 10, 15, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Adjust row heights
        ws.row_dimensions[1].height = 25  # Title row
        ws.row_dimensions[2].height = 20  # Subtitle row
        ws.row_dimensions[4].height = 20  # Header row

        # Create response
        filename = "Laporan_Harian.xlsx" if is_daily_report else "Laporan_Penjualan.xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response

    except Exception as e:
        print(f"Error exporting sales: {str(e)}")
        messages.error(request, f"Error exporting sales: {str(e)}")
        return redirect('simple_shop:sales')

@login_required
def member_search_page(request):
    """Render the member search page"""
    return render(request, 'members/search_member.html')

# Update the existing search_member view to handle both GET and POST
@login_required
def search_member(request):
    if request.method == 'GET':
        phone = request.GET.get('phone')
        try:
            member = Member.objects.get(phone=phone)
            return JsonResponse({
                'success': True,
                'member': {
                    'id': str(member.id),
                    'name': member.name,
                    'points': member.points,
                    'phone': member.phone
                }
            })
        except Member.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Member tidak ditemukan'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import User  # Adjust this import based on your actual model name
# ...existing imports...
# ...existing code...
def user_update(request, uuid):
    """
    View untuk memperbarui informasi pengguna
    - Admin hanya dapat mengubah status aktif/nonaktif pengguna lain
    - User hanya dapat mengubah data dirinya sendiri
    """
    User = get_user_model()
    user = get_object_or_404(User, pk=uuid)
    
    # Check if current user is editing themselves or is admin
    is_self = request.user.pk == user.pk
    is_admin_editing_other = (not is_self) and is_admin(request.user)
    
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        
        if form.is_valid():
            # If admin is editing another user, only allow changing is_active status
            if is_admin_editing_other:
                # Only update the is_active field and ignore other changes
                user.is_active = form.cleaned_data.get('is_active', user.is_active)
                user.save(update_fields=['is_active'])
                
                messages.success(request, f"Status pengguna {user.username} berhasil diperbarui")
            else:
                # Normal save for self-editing
                user_obj = form.save(commit=False)
                
                # Handle password change if provided
                if form.cleaned_data.get('password1'):
                    user_obj.set_password(form.cleaned_data['password1'])
                    password_changed = True
                else:
                    password_changed = False
                
                # Update role if applicable
                if form.cleaned_data.get('role'):
                    user_obj.role = form.cleaned_data['role']
                
                user_obj.save()
                
                if password_changed:
                    messages.success(request, "Informasi dan password pengguna berhasil diperbarui")
                    if request.user == user:
                        logout(request)
                        return redirect('simple_shop:login')
                else:
                    messages.success(request, "Informasi pengguna berhasil diperbarui")
            
            return redirect('simple_shop:users')
    else:
        form = UserForm(instance=user)
        form.fields['password1'].help_text = "Biarkan kosong jika tidak ingin mengubah password"
        form.fields['password2'].help_text = "Konfirmasi password baru"
    
    return render(request, 'users/user_update.html', {
        'form': form,
        'user': user,
        'is_admin_editing_other': is_admin_editing_other
    })

def forgot_password(request):
    """Handle forgot password requests"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        try:
            user = User.objects.get(email=email)
            
            # Generate a 6-digit OTP
            otp = ''.join(random.choices(string.digits, k=6))
            
            # Save OTP to session
            request.session['reset_otp'] = otp
            request.session['reset_email'] = email
            request.session['reset_expiry'] = (timezone.now() + timezone.timedelta(minutes=15)).isoformat()
            
            # Send OTP email with debug=False to ensure it uses SMTP
            send_otp_email(email, otp)
            
            messages.success(request, "Kode OTP telah dikirim ke email Anda. Silakan periksa kotak masuk Anda.")
            return render(request, 'registration/forgot_password.html', {
                'step': 'verify',
                'email': email
            })
            
        except User.DoesNotExist:
            messages.error(request, "Email tidak ditemukan dalam sistem kami.")
            
    return render(request, 'registration/forgot_password.html', {'step': 'email'})

def verify_otp(request):
    """Verify OTP for password reset"""
    if request.method == 'POST':
        email = request.POST.get('email')
        otp = request.POST.get('otp')
        # Debug info
        print(f"DEBUG - OTP Verification Request:")
        print(f"Email: {email}")
        print(f"Input OTP: {otp}")
        print(f"Session OTP: {request.session.get('reset_otp')}")
        print(f"Session Email: {request.session.get('reset_email')}")
        print(f"Session Expiry: {request.session.get('reset_expiry')}")
            
        stored_otp = request.session.get('reset_otp')
        stored_email = request.session.get('reset_email')
        expiry_str = request.session.get('reset_expiry')
        
        # Reset condition checking
        if not stored_otp or not stored_email or not expiry_str:
            messages.error(request, "Sesi OTP tidak ditemukan atau telah kedaluwarsa. Silakan coba lagi.")
            return redirect('simple_shop:forgot_password')
        
        # Check if OTP has expired
        try:
            expiry = timezone.datetime.fromisoformat(expiry_str)
            if timezone.now() > expiry:
                messages.error(request, "Kode OTP telah kedaluwarsa. Silakan minta kode baru.")
                return redirect('simple_shop:forgot_password')
        except Exception as e:
            messages.error(request, "Format waktu tidak valid. Silakan coba lagi.")
            return redirect('simple_shop:forgot_password')
        
        # Check if OTP and email are correct
        if otp == stored_otp and email == stored_email:
            return render(request, 'registration/forgot_password.html', {
                'step': 'reset',
                'email': email,
                'otp': otp
            })
        else:
            messages.error(request, "Kode OTP tidak valid. Silakan coba lagi.")
            return render(request, 'registration/forgot_password.html', {
                'step': 'verify',
                'email': email,
                'debug': settings.DEBUG
            })
    
    return redirect('simple_shop:forgot_password')

def resend_otp(request):
    """Resend OTP to user email"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        if User.objects.filter(email=email).exists():
            # Generate a new 6-digit OTP
            otp = ''.join(random.choices(string.digits, k=6))
            
            # Save OTP to session
            request.session['reset_otp'] = otp
            request.session['reset_email'] = email
            request.session['reset_expiry'] = (timezone.now() + timezone.timedelta(minutes=15)).isoformat()
            
            # Send OTP email
            send_otp_email(email, otp)
            
            messages.success(request, "Kode OTP baru telah dikirim ke email Anda.")
        else:
            messages.error(request, "Email tidak ditemukan dalam sistem kami.")
        
        return render(request, 'registration/forgot_password.html', {
            'step': 'verify',
            'email': email
        })
    
    return redirect('simple_shop:forgot_password')

def reset_password(request):
    """Reset user password after OTP verification"""
    if request.method == 'POST':
        email = request.POST.get('email')
        otp = request.POST.get('otp')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        stored_otp = request.session.get('reset_otp')
        stored_email = request.session.get('reset_email')
        expiry_str = request.session.get('reset_expiry')
        
        # Validate session
        if not all([stored_otp, stored_email, expiry_str]):
            messages.error(request, "Sesi reset password tidak valid atau telah kedaluwarsa.")
            return redirect('simple_shop:forgot_password')
        
        # Check expiry
        expiry = timezone.datetime.fromisoformat(expiry_str)
        if timezone.now() > expiry:
            messages.error(request, "Sesi reset password telah kedaluwarsa. Silakan coba lagi.")
            return redirect('simple_shop:forgot_password')
        
        # Verify OTP and email
        if otp != stored_otp or email != stored_email:
            messages.error(request, "Data verifikasi tidak valid. Silakan coba lagi.")
            return redirect('simple_shop:forgot_password')
        
        # Validate password
        if new_password != confirm_password:
            messages.error(request, "Password dan konfirmasi password tidak cocok.")
            return render(request, 'registration/forgot_password.html', {
                'step': 'reset',
                'email': email,
                'otp': otp
            })
        
        # Validate password complexity
        if len(new_password) < 8:
            messages.error(request, "Password harus memiliki minimal 8 karakter.")
            return render(request, 'registration/forgot_password.html', {
                'step': 'reset',
                'email': email,
                'otp': otp
            })
        
        try:
            # Update user's password
            user = User.objects.get(email=email)
            user.set_password(new_password)
            user.save()
            
            # Clear session data
            for key in ['reset_otp', 'reset_email', 'reset_expiry']:
                if key in request.session:
                    del request.session[key]
            
            messages.success(request, "Password berhasil diubah. Silakan login dengan password baru Anda.")
            return redirect('simple_shop:login')
        except User.DoesNotExist:
            messages.error(request, "Terjadi kesalahan. Pengguna tidak ditemukan.")
            return redirect('simple_shop:forgot_password')
    
    return redirect('simple_shop:forgot_password')

# Perbaiki fungsi send_otp_email di bagian bawah file
# Tambahkan impor untuk email_utils
from .email_utils import send_otp_email as send_otp_mail

# Ganti fungsi send_otp_email yang bermasalah
def send_otp_email(email, otp):
    """Helper function to send OTP email"""
    return send_otp_mail(email, otp, debug=True)

from django.shortcuts import render
# ...other imports...

# ...other view functions...

@login_required
def debug_sales(request):
    """A debug view to help diagnose sales data issues"""
    if not settings.DEBUG:
        raise Http404("Debug view only available in DEBUG mode")
    
    # Get a sample of sales data for debugging
    sales_qs = Sale.objects.select_related('user', 'member', 'payment').order_by('-created_at')
    sales_list = list(sales_qs[:10].values('id', 'created_at', 'total_price', 'user__username', 'member__name'))
    
    # Convert to JSON for display
    sales_json = json.dumps(sales_list, cls=DjangoJSONEncoder, indent=2)
    
    context = {
        'sales': sales_qs[:10],
        'sales_debug': sales_json,
        'search_query': request.GET.get('search', ''),
        'date_from': request.GET.get('date_from'),
        'date_to': request.GET.get('date_to'),
        'payment_method': request.GET.get('payment_method')
    }
    
    return render(request, 'sales/debug_sales.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Member

@login_required
def member_search_page(request):
    return render(request, 'members/search_member.html')

@login_required
def member_search_api(request):
    phone = request.GET.get('phone', '').strip()
    print(f"Search API called with phone: {phone}")  # Debug log
    
    try:
        member = Member.objects.filter(phone=phone).first()
        print(f"Search result: {member}")  # Debug log
        
        if member:
            # Make sure to include the ID in the response
            member_data = {
                'id': str(member.id),  # Convert UUID to string
                'name': member.name,
                'phone': member.phone,
                'points': member.points or 0
            }
            print(f"Sending member data: {member_data}")  # Debug log
            
            return JsonResponse({
                'success': True,
                'member': member_data
            })
            
        return JsonResponse({
            'success': False,
            'status': 'not_found',
            'phone': phone,
            'message': f'Member dengan nomor {phone} tidak ditemukan'
        })
    except Exception as e:
        print(f"Error in search: {str(e)}")  # Debug log
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@login_required
def member_register_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name')
            phone = data.get('phone')
            
            if not name or not phone:
                return JsonResponse({
                    'success': False,
                    'message': 'Nama dan nomor telepon harus diisi'
                })
            
            # Check if member already exists using get_or_create
            customer_type = CustomerType.objects.get_or_create(name='Regular')[0]
            member, created = Member.objects.get_or_create(
                phone=phone,
                defaults={
                    'name': name,
                    'customer_type': customer_type,
                    'points': 0
                }
            )
            
            if not created:
                # Member already exists, return existing data
                return JsonResponse({
                    'success': True,
                    'member': {
                        'id': str(member.id),
                        'name': member.name,
                        'phone': member.phone,
                        'points': member.points
                    },
                    'message': 'Member dengan nomor ini sudah terdaftar'
                })
            
            # New member created successfully
            return JsonResponse({
                'success': True,
                'member': {
                    'id': str(member.id),
                    'name': member.name,
                    'phone': member.phone,
                    'points': member.points
                },
                'message': f'Member {name} berhasil didaftarkan'
            })
            
        except Exception as e:
            print(f"Error registering member: {str(e)}")  # Debug log
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'message': 'Method tidak diizinkan'
    })

@login_required
def member_add(request):
    # Implementation for adding member manually
    pass

@login_required
def member_edit(request, pk):
    # Implementation for editing member
    pass

@login_required
def print_member_card(request, member_id):
    """Print a member card PDF with clean, elegant design and balanced layout"""
    try:
        print(f"Generating streamlined member card for ID: {member_id}")
        
        # Explicitly handle UUID validation
        if not isinstance(member_id, uuid.UUID):
            try:
                member_id = uuid.UUID(str(member_id))
            except ValueError:
                messages.error(request, 'ID Member tidak valid')
                return redirect('simple_shop:member_search_page')
        
        member = get_object_or_404(Member, id=member_id)
        
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm, inch
        from reportlab.lib.colors import Color, HexColor
        from reportlab.lib.utils import ImageReader
        import qrcode
        from io import BytesIO
        
        # Set up PDF dimensions (credit card size)
        width, height = 85.6*mm, 53.98*mm
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=(width, height))
        
        # ===== ELEGANT COLOR SCHEME =====
        primary_color = HexColor('#1a237e')     # Deep blue for background
        accent_color = HexColor('#42a5f5')      # Light blue for accents
        highlight_color = HexColor('#ffeb3b')   # Yellow for highlights
        dark_accent = HexColor('#0d47a1')       # Dark blue for top bar
        
        # ===== CARD BACKGROUND =====
        # Main background
        c.setFillColor(primary_color)
        c.rect(0, 0, width, height, fill=1)
        
        # Top accent bar - exactly 12mm height
        c.setFillColor(dark_accent)
        c.rect(0, height-12*mm, width, 12*mm, fill=1)
        
        # Left accent stripe - precisely 2mm wide
        c.setFillColor(accent_color)
        c.rect(0, 0, 2*mm, height, fill=1)
        
        # Add subtle pattern to background (smaller, more elegant dots)
        c.setFillColor(HexColor('#FFFFFF'))
        c.setFillAlpha(0.03)
        for i in range(20, int(width), 10):
            for j in range(5, int(height), 10):
                c.circle(i, j, 0.2, fill=1)
        c.setFillAlpha(1)
        
        # ===== CARD HEADER WITH PERFECT ALIGNMENT =====
        # Brand logo area
        c.setFillColor(HexColor('#FFFFFF'))
        c.setFont("Helvetica-Bold", 16)
        c.drawString(7.5*mm, height-8*mm, "IQMart")
        
        # Tagline
        c.setFillColor(HexColor('#FFFFFFAA'))
        c.setFont("Helvetica-Oblique", 5.5)
        c.drawString(7.5*mm, height-11*mm, "SMART SHOPPING EXPERIENCE")
        
        # Member text without badge background
        c.setFillColor(HexColor('#FFFFFF'))  # White text instead of black on yellow
        c.setFont("Helvetica-Bold", 6.5)
        c.drawRightString(width-5*mm, height-8*mm, "MEMBER CARD")
        
        # Clean divider line - perfectly positioned
        c.setStrokeColor(HexColor('#FFFFFF33'))
        c.setLineWidth(0.3)
        c.line(7.5*mm, height-15*mm, width-7.5*mm, height-15*mm)
        
        # ===== MEMBER INFORMATION SECTION - PERFECTLY SPACED =====
        # Create consistent spacing for all information
        info_left_margin = 7.5*mm
        info_section_start = height-20*mm
        section_spacing = 8*mm
        
        # Name section
        c.setFillColor(accent_color)
        c.setFont("Helvetica", 6)
        c.drawString(info_left_margin, info_section_start, "NAME")
        
        # Name value with proper vertical spacing
        c.setFillColor(HexColor('#FFFFFF'))
        c.setFont("Helvetica-Bold", 11)
        name = member.name
        if len(name) > 20:
            name = name[:17] + "..."
        c.drawString(info_left_margin, info_section_start-5*mm, name)
        
        # Phone section with proper spacing
        c.setFillColor(accent_color)
        c.setFont("Helvetica", 6)
        c.drawString(info_left_margin, info_section_start-section_spacing, "PHONE")
        
        # Phone value
        c.setFillColor(HexColor('#FFFFFF'))
        c.setFont("Helvetica", 9)
        c.drawString(info_left_margin, info_section_start-section_spacing-5*mm, member.phone)
        
        # ===== QR CODE SECTION - PERFECTLY ALIGNED =====
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=0,
        )
        qr.add_data(f"ID:{member.id}|PHONE:{member.phone}")
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format="PNG")
        qr_buffer.seek(0)
        
        # QR code with perfect positioning - moved slightly higher for better spacing
        qr_width, qr_height = 18*mm, 18*mm
        qr_x = width - qr_width - 7*mm
        qr_y = height - 36*mm  # Moved up by 1mm to create more space below
        
        # White frame for QR code - precise measurements
        frame_padding = 1.5*mm
        c.setFillColor(HexColor('#FFFFFF'))
        c.roundRect(
            qr_x - frame_padding/2, 
            qr_y - frame_padding/2, 
            qr_width + frame_padding, 
            qr_height + frame_padding, 
            2*mm, 
            fill=1
        )
        
        # Draw QR code with perfect alignment
        c.drawImage(ImageReader(qr_buffer), qr_x, qr_y, width=qr_width, height=qr_height)
        
        # ===== ISSUE DATE AND VALIDITY SECTION - LEFT AND RIGHT ALIGNED =====
        # Increased vertical distance from QR code
        validity_line_y = 13*mm  # Moved down to create more space from QR code
        
        # Issue date text at left side
        if hasattr(member, 'created_at') and member.created_at:
            reg_date = member.created_at.strftime('%d/%m/%Y')
        else:
            reg_date = "N/A"
        
        # Left side - Issued date
        c.setFillColor(HexColor('#FFFFFF99'))
        c.setFont("Helvetica", 5)
        c.drawString(7.5*mm, validity_line_y, f"ISSUED: {reg_date}")
        
        # Right side - VALID PERMANENTLY
        c.setFillColor(HexColor('#FFFFFF99'))
        c.setFont("Helvetica", 5)
        c.drawRightString(width-7.5*mm, validity_line_y, "VALID PERMANENTLY")
        
        # ===== MORE SUBTLE BLACK FOOTER =====
        # Footer background - more subtle black (semi-transparent)
        footer_height = 9*mm
        c.setFillColor(HexColor('#000000'))
        c.setFillAlpha(0.85)  # Make it slightly translucent for subtlety
        c.rect(0, 0, width, footer_height, fill=1)
        c.setFillAlpha(1)  # Reset fill alpha
        
        # Add a subtle horizontal line separator above footer
        c.setStrokeColor(HexColor('#FFFFFF33'))
        c.setLineWidth(0.2)
        c.line(7.5*mm, footer_height + 1*mm, width-7.5*mm, footer_height + 1*mm)
        
        # ID text - truncate if too long
        id_text = str(member.id)
        footer_vertical_center = 4.5*mm
        
        # ID on left with more spacing for readability
        c.setFillColor(highlight_color)
        c.setFont("Helvetica-Bold", 5.5)
        c.drawString(5*mm, footer_vertical_center, "ID:")
        
        # ID value with adjusted spacing
        c.setFillColor(HexColor('#FFFFFF'))
        if len(id_text) > 15:
            # For long IDs, use smaller font and truncate if needed
            c.setFont("Helvetica", 4)
            id_text = id_text[:18] + "..." if len(id_text) > 21 else id_text
            c.drawString(10*mm, footer_vertical_center, id_text)
        else:
            c.setFont("Helvetica", 5)
            c.drawString(10*mm, footer_vertical_center, id_text)
        
        # Website at center with proper spacing - moved further right
        # This ensures it doesn't crowd the ID text
        c.setFillColor(HexColor('#FFFFFF'))
        c.setFont("Helvetica", 6)
        # Positioning the website text more to the right side for better spacing
        c.drawCentredString(width*0.75, footer_vertical_center, "www.iqmart.id")
        
        # Finish the PDF
        c.save()
        buffer.seek(0)
        
        # Return PDF file
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="member_card_{member.id}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        
        return response
        
    except Exception as e:
        print(f"Error in print_member_card: {str(e)}")
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'Error generating member card: {str(e)}')
        return redirect('simple_shop:member_search_page')


@login_required
@user_passes_test(is_admin)
def toggle_user_status(request):
    """
    AJAX view to toggle a user's active status directly from the user list
    Also updates staff status when deactivating a user
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('user_id')
            
            # Get the user
            User = get_user_model()
            user = get_object_or_404(User, id=user_id)
            
            # Don't allow changing your own status
            if user == request.user:
                return JsonResponse({
                    'success': False,
                    'message': 'Anda tidak dapat mengubah status akun Anda sendiri'
                }, status=403)
            
            # Toggle the status
            user.is_active = not user.is_active
            
            # When deactivating a user, also remove staff status
            # When activating, restore staff status if they have a role
            if not user.is_active:
                user.is_staff = False
            elif user.role:  # Only restore staff status if they have a role
                user.is_staff = True
                
            # Save both changes in one go
            user.save(update_fields=['is_active', 'is_staff'])
            
            status_text = 'aktif' if user.is_active else 'nonaktif'
            staff_status = user.is_staff
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'is_staff': staff_status,
                'message': f'Status pengguna {user.username} berhasil diubah menjadi {status_text}'
            })
            
        except Exception as e:
            print(f"Error toggling user status: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Metode tidak diizinkan'
    }, status=405)


@login_required
def profile(request):
    """
    Redirect to user update page for the logged-in user
    This allows users to edit their own profile
    """
    return redirect('simple_shop:user_update', uuid=request.user.pk)


